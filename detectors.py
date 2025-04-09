import openpyxl as xl
from openpyxl import Workbook as wb
from abc import ABC, abstractmethod

class Test_Suite():
    __slots__ = ["wb_path", "wb", "to_run", "results"]

    def __init__(self, wb_path, to_run = None):
        self.wb_path = wb_path
        self.wb = xl.load_workbook(
            wb_path,
            keep_vba = False,
            rich_text = True,
            data_only = False
        )



        if to_run == None:
            self.to_run = [Grid_Edges, Header]
        else:
            self.to_run = to_run
        self.results = dict()


    def run(self):
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            self.results[sheet] = dict()

            queue = self.to_run[:]
            completed_tests = dict()
            while len(queue) > 0:
                    
                test_type = queue.pop(0)
                test = test_type(ws)
                
                if issubclass(test_type, Has_Dependency):

                    if any( 
                        map(
                            lambda depend: depend(ws).name not in completed_tests.keys(),
                            test.dependencies
                        ) 
                    ):
                        queue.append(test)
                        continue

                    else:
                        fulfilled = {dep_name : dep_test for dep_name, dep_test in completed_tests.items() if type(dep_test) in test.dependencies}
                        test.validate(**fulfilled)
                else:
                    test.validate()
                
                self.results[sheet][test.name] = test
                completed_tests[test.name] = test
                    



            
    
    def report(self):
        for sheet, tests in self.results.items():
            print(f"Sheet: {sheet}")
            for test_name, test in tests.items():
                test.report()
                print("\n")
            print("\n\n")




class Test(ABC):


    def __init__(self, ws, name):
        self.name = name
        self.is_run = False
        self.status = None
        self.ws = ws
        self.issues = dict()
        self.message = "Not yet run"

    @abstractmethod
    def validate(self):
        pass

    def report(self):
        print(f"name: {self.name}\nstatus: {self.status}\nmessage: {self.message}\nissues: {self.issues}")



class Feed_Forward():
    def __init__(self):
        self.output = dict()


class Has_Dependency():
    def __init__(self, *dependencies):
        self.dependencies = dependencies

    def check_input(self, *inputs):
        self.inputs = inputs

        for test in self.inputs:
            assert test.is_run, "Dependency not yet run"


class Grid_Edges(Test, Feed_Forward):

    def __init__(self, ws):
        Test.__init__(self, ws, "grid_edges")
        Feed_Forward.__init__(self)

    def validate(self):
        # Iterate through columns
        for idx, col in enumerate(self.ws.columns):


            # Check whether this entire column is empty
            # map visits each cell and checks whether contents empty
            # all checks whether every single cell empty
            if all( map(lambda cell : cell.internal_value == None, col)   ):

                # Keep looking for the first nonempty
                continue

            else:   

                # This column has nonempty cells, and it is the first
                # Print the index of this col
                self.output["col_start"] = idx + 1
                break

        # Iterate through rows
        for idx, row in enumerate(self.ws.rows):

            # Check whether this entire column is empty
            # map applies lambda func to each cell
            # lambda func checks whether contents empty 
            # all checks whether every single cell empty
            if all( map(lambda cell : cell.internal_value == None, row)   ):

                # Keep looking for the first nonempty
                continue

            else:   

                # This column has nonempty cells, and it is the first
                # Print the index of this row
                self.output["row_start"] = idx + 1

                # End the loop
                break

        if self.output["col_start"] == 1 and self.output["row_start"] == 1:
            self.status = True
            self.issues = []
            self.message = "Normal table position"
        else:
            self.status = False
            self.issues = self.output
            self.message = "Blank space before first column or row"
        
        self.is_run = True


class Header(Test, Has_Dependency):
    __slots__ = ["row_start"]

    def __init__(self, ws):
        Test.__init__(self, ws, "headers")
        Has_Dependency.__init__(self, Grid_Edges)
        self.row_start = 1
            

    def validate(self, grid_edges):
        self.check_input(grid_edges)
        self.row_start = grid_edges.output["row_start"]

        for col in self.ws.iter_cols( min_row = self.row_start, max_row = self.row_start, min_col = grid_edges.output["col_start"]):
            
            cell = col[0]
            header = cell.internal_value
            if self.is_valid(header):
                continue
            else:
                self.issues[cell.coordinate] = f"invalid header: {header}"

        self.is_run = True
        if self.issues:
            self.status = False
            self.message = "Headers are not acceptable"
        else:
            self.status = True
            self.message = "Headers meet expectations"


    def is_valid(self, header):
        validity = isinstance(header, str)
        validity = validity and (len(header) >= 3 and len(header) <= 12)
        validity = validity and not header[0].isdigit()
        validity = validity and not (" " in header)
        return validity
    

if __name__ == "__main__":
    suite = Test_Suite("data/demo.xlsx")
    suite.run()
    suite.report()